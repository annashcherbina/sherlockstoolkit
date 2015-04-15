from openpyxl import Workbook,load_workbook
import Params; 
import matplotlib 
matplotlib.use('Agg')
import matplotlib.pyplot as plt 
from pylab import stem; 
from pylab import setp; 
import numpy as np; 

def filter_snp(maf,countall,forward_count): 
    #filter strand-biased/ low count SNPs 
    print "countall:"+str(countall) 
    print "forward_count:"+str(forward_count) 
    print "maf:"+str(maf) 
    if forward_count < Params.min_calls: 
        return True,'low' 
    if (countall-forward_count) < Params.min_calls: 
        return True,'low'

    #filter ambiguous SNPs 
    if (maf > Params.maf_thresh_individual_low) and (maf < Params.maf_thresh_individual_amb): 
        print "in filter_snp:" 
        print "Params.maf_thresh_individual_low:"+str(Params.maf_thresh_individual_low) 
        print "Params.maf_thresh_individual_amb:"+str(Params.maf_thresh_individual_amb) 
        return True ,'ambiguous'
    if (maf > Params.maf_thresh_individual_amb_upper) and (maf < Params.maf_thresh_individual_high): 
        print "in filter snp:"
        print "Params.maf_thresh_individual_amb_upper:"+str(Params.maf_thresh_individual_amb_upper) 
        print "Params.maf_thresh_individual_high:"+str(Params.maf_thresh_individual_high) 
        return True ,'ambiguous'
    return False ,'good'

def outputExcelSummary(maf_dict,ma_snps,replicates_global,truth_sample,internal_hash,metadata,cur,user_id,folder_id,parameter_group,quality,locus_group,locus_group_snps):
    replicates=[]
    for item in replicates_global: 
        replicates.append(item) 
    attachment_id=0 
    wb = Workbook()
    ws=wb.get_active_sheet()
    ws.title='All mA' 
    #print "metadata:"+str(metadata) 
    #Add the header row to the output Excel File 
    c=ws.cell(row=0,column=0) 
    c.value="Locus"
    startx=1
    endx=0
    if truth_sample: 
        c=ws.cell(row=0,column=1) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference)"
        c=ws.cell(row=0,column=2) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference) Count"
        c=ws.cell(row=0,column=3) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference) MAF"
        startx=4
        endx=1
    #get file_name that the user will see 
    title="ReplicateDetails"
    panels=set([]) 
    sources=set([]) 
    if truth_sample: 
        title=title+"_Reference_"+metadata[truth_sample][0]+"_"+metadata[truth_sample][1] 
        panels.add(metadata[truth_sample][-2])
        sources=sources.union(metadata[truth_sample][-1]) 
    for i in range(len(replicates)): 
        if replicates[i]!=truth_sample: 
            #print str(metadata[replicates[i]])
            title=title+"_"+(metadata[replicates[i]])[0]+"_"+(metadata[replicates[i]])[1]
            panels.add(metadata[replicates[i]][-2])
            print str(metadata[replicates[i]][-1])
            sources=sources.union(metadata[replicates[i]][-1]) 
    if len(title)>250: 
        title="ReplicateDetails_"+str(len(replicates))
        for panel in panels: 
            title=title+"_"+panel
        for source in sources: 
            title=title+"_"+source
    title=title+".xlsx"
    print str(metadata) 
    #Populate the contributors and description fields 
    description="Quality:"+str(quality)+"\n"+"Locus Group:"+str(locus_group)+"\n"
    if truth_sample:         
        description=description+"Reference:\n"+metadata[truth_sample][3]+" ("+metadata[truth_sample][0]+")"+ " "+metadata[truth_sample][1] +" ("+metadata[truth_sample][2]+")" +"\n" 
    description=description+"Replicates:"
    for rep in replicates: 
        if rep==truth_sample: 
            continue 
        description=description+"\n"+metadata[rep][3]+" (" + metadata[rep][0]+")"+ " "+metadata[rep][1]+" ("+metadata[rep][2]+")"
    if truth_sample in replicates: 
        replicates.remove(truth_sample) 

    #Write the rest of the header row to Excel output file 
    for j in range(startx,3*len(replicates)+1+endx,3): 
        c=ws.cell(row=0,column=j) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1] 
        c=ws.cell(row=0,column=j+1) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1]+ " Count" 
        c=ws.cell(row=0,column=j+2) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1]+ " MAF" 
        
    discrepancy_snps=[] 
    #write values to the Excel file 
    for i in range(1,len(ma_snps)): 
        snp_name=ma_snps[i-1] 
        c=ws.cell(row=i,column=0)
        c.value=snp_name
        startx=1
        endx=0 
        all_values=set([]) 
        if truth_sample: 
            c=ws.cell(row=i,column=1) 
            if truth_sample not in maf_dict[snp_name]: 
                c.value="NOP" 
                all_values.add("NOP") 
                maf=0 
                count=0
            else: 
                maf,count,forward_count=maf_dict[snp_name][truth_sample]
                bad,reason=filter_snp(maf,count,forward_count)
                print "bad:"+str(bad)
                print "reason:"+str(reason) 
                if (bad ==True) and (reason=='low'): 
                    c.value="NC"
                    all_values.add("NC") 
                elif (bad==True) and (reason=='ambiguous'): 
                    c.value="AMB" 
                    all_values.add("AMB") 
                elif maf < Params.maf_thresh_individual_low: 
                    c.value=0 
                    all_values.add(0) 
                elif maf > (Params.maf_thresh_individual_high): 
                    c.value=2 
                    all_values.add(2) 
                elif (maf > Params.maf_thresh_individual_amb) and (maf < Params.maf_thresh_individual_amb_upper): 
                    c.value=1 
                    all_values.add(1) 
            c=ws.cell(row=i,column=2) 
            c.value=count 
            c=ws.cell(row=i,column=3) 
            c.value=maf 
            startx=4
            endx=1 
        for j in range(startx,3*len(replicates)+1+endx,3): 
            c=ws.cell(row=i,column=j) 
            rep=replicates[(startx-j)/3]
            if rep not in maf_dict[snp_name]: 
                c.value="NOP"
                all_values.add("NOP") 
                maf=0
                count=0 
            else: 
                maf,count,forward_count=maf_dict[snp_name][rep] 
                bad,reason=filter_snp(maf,count,forward_count)
                print "bad:"+str(bad)
                print "reason:"+str(reason) 
                if (bad==True) and (reason=='low'):                     
                    c.value="NC" 
                    all_values.add("NC") 
                elif (bad==True) and (reason=='ambiguous'): 
                    c.value="AMB" 
                    all_values.add("AMB") 
                elif maf < Params.maf_thresh_individual_low: 
                    c.value=0 
                    all_values.add(0) 
                elif maf > (Params.maf_thresh_individual_high): 
                    c.value=2 
                    all_values.add(2) 
                elif ((maf > Params.maf_thresh_individual_amb) and (maf < Params.maf_thresh_individual_amb_upper)): 
                    c.value=1 
                    all_values.add(1) 
            c=ws.cell(row=i,column=j+1) 
            c.value=count 
            c=ws.cell(row=i,column=j+2) 
            c.value=maf 
        #print "all_values:"+str(all_values)     
        if len(all_values) > 1: 
            discrepancy_snps.append(snp_name) 
    #print "discrepancy snps:"+str(discrepancy_snps) 
    #create another sheet in the output excel file to store info about discrepancies. 
    ws2=wb.create_sheet() 
    ws2.title='Discrepancies' 
    c=ws2.cell(row=0,column=0) 
    c.value="Locus" 
    startx=1
    if truth_sample: 
        c=ws2.cell(row=0,column=1) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference)"
        c=ws2.cell(row=0,column=2) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference) Count"
        c=ws2.cell(row=0,column=3) 
        c.value=metadata[truth_sample][0]+"_"+metadata[truth_sample][1]+"(Reference) MAF"
        startx=4
    #Write the rest of the header row to Excel output file 
    for j in range(startx,3*len(replicates)+1,3): 
        c=ws2.cell(row=0,column=j) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1] 
        c=ws2.cell(row=0,column=j+1) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1]+ " Count" 
        c=ws2.cell(row=0,column=j+2) 
        c.value=metadata[replicates[(j-startx)/3]][0]+"_"+metadata[replicates[(j-startx)/3]][1]+ " MAF" 
    for i in range(1,len(discrepancy_snps)+1): 
        snp_name=discrepancy_snps[i-1] 
        c=ws2.cell(row=i,column=0) 
        c.value=snp_name
        if truth_sample: 
            c=ws2.cell(row=i,column=1) 
            if truth_sample not in maf_dict[snp_name]: 
                c.value="NOP"
                maf=0
                count=0 
            else: 
                maf,count,forward_count=maf_dict[snp_name][truth_sample]
                bad,reason=filter_snp(maf,count,forward_count) 
                if (bad==True) and (reason=='low'): 
                    c.value="NC"
                if (bad==True) and (reason=='ambiguous'): 
                    c.value="AMB" 
                elif maf < Params.maf_thresh_individual_low: 
                    c.value=0 
                elif maf > (Params.maf_thresh_individual_high): 
                    c.value=2 
                elif ((maf > Params.maf_thresh_individual_amb) and (maf < Params.maf_thresh_individual_amb_upper)): 
                    c.value=1 
            c=ws2.cell(row=i,column=2) 
            c.value=count 
            c=ws2.cell(row=i,column=3) 
            c.value=maf 
        for j in range(startx,3*len(replicates)+1+endx,3): 
            c=ws2.cell(row=i,column=j) 
            rep=replicates[(startx-j)/3] 
            if rep not in maf_dict[snp_name]: 
                c.value="NOP" 
                maf=0; 
                count=0 
            else: 
                maf,count,forward_count=maf_dict[snp_name][rep] 
                bad,reason=filter_snp(maf,count,forward_count) 
                if (bad==True) and (reason=='low'): 
                    c.value="NC" 
                elif (bad==True) and (reason=='ambiguous'): 
                    c.value="AMB" 
                elif maf < Params.maf_thresh_individual_low: 
                    c.value=0 
                elif maf > Params.maf_thresh_individual_high: 
                    c.value=2 
                elif ((maf > Params.maf_thresh_individual_amb) and (maf < Params.maf_thresh_individual_amb_upper)): 
                    c.value=1 
            c=ws2.cell(row=i,column=j+1) 
            c.value=count 
            c=ws2.cell(row=i,column=j+2) 
            c.value=maf 
            
    ws3=wb.create_sheet() 
    query="select name from users where id="+str(user_id)+";" 
    cur.execute(query) 
    username=cur.fetchone()
    if username!=None: 
        username=username[0] 
    ws3.title="Parameters"
    c=ws3.cell(row=0,column=0)
    c.value= str(parameter_group)+","+str(username)   
    c=ws3.cell(row=1,column=0) 
    c.value="Minimum Reads Per Locus" 
    c=ws3.cell(row=1,column=1) 
    c.value=Params.min_calls 

    c=ws3.cell(row=2,column=0) 
    c.value="Reference - MAF for 0 Minor Alleles (upper bound)"
    c=ws3.cell(row=2,column=1) 
    c.value=Params.maf_thresh_individual_low

    c=ws3.cell(row=3,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (lower bound)"
    c=ws3.cell(row=3,column=1) 
    c.value=Params.maf_thresh_individual_amb


    c=ws3.cell(row=4,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (upper bound)"
    c=ws3.cell(row=4,column=1) 
    c.value=Params.maf_thresh_individual_amb_upper

    c=ws3.cell(row=5,column=0) 
    c.value="Reference - MAF for 2 Minor Alleles (lower bound)"
    c=ws3.cell(row=5,column=1) 
    c.value=Params.maf_thresh_individual_high

    c=ws3.cell(row=6,column=0) 
    c.value="Strand Bias Cumulative Count (lower bound)"
    c=ws3.cell(row=6,column=1) 
    c.value=Params.strand_bias_thresh

    c=ws3.cell(row=7,column=0) 
    c.value="Ambiguous Bias Cumulative Count (lower bound)"
    c=ws3.cell(row=7,column=1) 
    c.value=Params.ambiguous_thresh 

    
    c=ws3.cell(row=8,column=0) 
    c.value="Quality" 
    c=ws3.cell(row=8,column=1) 
    c.value=quality 

    c=ws3.cell(row=9,column=0) 
    c.value="Locus Group" 
    c=ws3.cell(row=9,column=1) 
    c.value=locus_group

    if len(locus_group_snps)>0: 
        #add a worksheet storing the locus group snps 
        ws4=wb.create_sheet()         
        ws4.title="Locus Group SNPs" 
        c=ws4.cell(row=0,column=0)
        c.value="Locus Group" 
        c=ws4.cell(row=0,column=1) 
        c.value=locus_group
        locus_group_snps_string='|'.join([str(i) for i in locus_group_snps]); 
        query="select name from loci where id REGEXP '("+locus_group_snps_string+")';" 
        cur.execute(query) 
        locus_group_snp_names=[i[0] for i in cur.fetchall()] 
        for i in range(len(locus_group_snp_names)): 
            c=ws4.cell(row=i+1,column=0) 
            c.value=locus_group_snp_names[i] 
    
    wb.save(Params.output_details_excel+title) 
    #Insert output excel file and metadata into the database 
    #param_hash=str(Params.maf_thresh_individual_low_id)+str(Params.maf_thresh_individual_amb_id)+str(Params.min_calls_id)+str(Params.ambiguous_thresh_id)+str(Params.strand_bias_thresh_id) 
    #internal_hash=hash(param_hash+internal_hash) 
    query="INSERT into attachments (user_id,folder_id,file_name,internal_hash,description,file_type,content_type,updated_at,codebase_version,contents) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+"',"+str(internal_hash)+",'"+str(description)+"','Replicate Details','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=(open(Params.output_details_excel+title,'r').read(),) 
    cur.execute(query,args)
    query="select id from attachments where internal_hash='"+str(internal_hash)+"';"
    cur.execute(query) 
    attachment_id=cur.fetchone()[0] 
    return attachment_id 

def outputDetailsImage(maf_dict,ma_snps,replicates_global,truth_sample,internal_hash,metadata,cur,user_id,folder_id,snp_to_pos,quality,locus_group): 
    replicates=[]
    for item in replicates_global: 
        replicates.append(item) 


    #get file_name that the user will see 
    panels=set([]) 
    sources=set([]) 
    title="ReplicateDetails"
    if truth_sample: 
        title=title+"_Reference_"+metadata[truth_sample][0]+"_"+metadata[truth_sample][1] 
        sources=sources.union(metadata[truth_sample][-1]) 
        panels.add(metadata[truth_sample][-2]) 
    for i in range(len(replicates)): 
        if replicates[i]!=truth_sample: 
            title=title+"_"+(metadata[replicates[i]])[0]+"_"+(metadata[replicates[i]])[1]
            sources=sources.union(metadata[replicates[i]][-1]) 
            panels.add(metadata[replicates[i]][-2]) 
    if len(title) > 250: 
        title="ReplicateDetails_"+str(len(replicates))
        for panel in panels: 
            title=title+"_"+panel 
        for source in sources: 
            title=title+"_"+source 
    title=title+".png"

    #Populate the contributors and description fields 
    description="Quality:"+str(quality)+"\n"+"Locus Group:"+str(locus_group)+"\n" 
    if truth_sample: 
        description=description+"Reference:\n"+metadata[truth_sample][3]+" ("+metadata[truth_sample][0]+")"+ " "+metadata[truth_sample][1] +" ("+metadata[truth_sample][2]+")" +"\n" 
    description=description+"Replicates:"
    for rep in replicates: 
        if rep==truth_sample: 
            continue 
        description=description+"\n"+metadata[rep][3]+" (" + metadata[rep][0]+")"+ " "+metadata[rep][1]+" ("+metadata[rep][2]+")"

    details_image_id=0 
    numreps=len(replicates)
    #print "Numreps:"+str(len(replicates))
    legend_names=['True Pos','No Call','False Pos','False Neg', 'Not on Panel']
    if truth_sample: 
        plt.subplot(numreps,1,1) 
        xvals=[] 
        yvals=[] 
        for snp in ma_snps: 
            if truth_sample not in maf_dict[snp]: 
                continue 
            maf,countall,forward_count=maf_dict[snp][truth_sample] 
            bad,reason=filter_snp(maf,countall,forward_count) 
            if bad:
                continue 
            print "truthmaf:"+str(maf) 
            print "maf_thresh_individual_high:"+str(Params.maf_thresh_individual_high) 
            print "maf_thresh_individual_low:"+str(Params.maf_thresh_individual_low) 
            if maf > (Params.maf_thresh_individual_high): 
                yvals.append(2) 
                xvals.append(snp_to_pos[snp]) 
            elif maf > (Params.maf_thresh_individual_low): 
                yvals.append(1) 
                xvals.append(snp_to_pos[snp])
        if len(xvals)>0 and len(yvals)>0:
            markerline, stemlines, baseline = stem(xvals, yvals, '-')
            setp(stemlines, 'color', 'b','linewidth',1.5)
            setp(markerline,'marker','None') 
        plt.xlabel('SNP') 
        plt.ylabel('Minor Alleles') 
        plt.title('Reference: '+str(metadata[truth_sample][0])+" "+str(metadata[truth_sample][1]))
        if truth_sample in replicates: 
            replicates.remove(truth_sample) 
    print "truth_sample:"+str(truth_sample) 
    for i in range(len(replicates)): 
        nc_x=[] 
        nc_y=[] 
        fn_x=[] 
        fn_y=[] 
        fp_x=[] 
        fp_y=[] 
        nop_x=[] 
        nop_y=[] 
        xvals=[] 
        yvals=[] 

        plots = []
        proxies = []

        if truth_sample: 
            index=i+2
        else: 
            index=i+1
        plt.subplot(numreps,1,index) 
        for snp in ma_snps:
            if (truth_sample !=None) and (truth_sample  not in maf_dict[snp]): 
                continue 
            if replicates[i] not in maf_dict[snp]: 
                nop_x.append(snp_to_pos[snp]) 
                nop_y.append(2) 
            else: 
                maf,count,forward_count=maf_dict[snp][replicates[i]] 
                bad,reason=filter_snp(maf,count,forward_count) 
                if (bad==True) and (reason=='low'): 
                    nc_x.append(snp_to_pos[snp]) 
                    nc_y.append(2) 
                else:
                    #is this a false negative? 
                    if truth_sample and (maf < Params.maf_thresh_individual_low) and (maf_dict[snp][truth_sample][1] > Params.min_calls) and (maf_dict[snp][truth_sample][0] > Params.maf_thresh_individual_low): 
                        fn_x.append(snp_to_pos[snp]) 
                        if maf > (Params.maf_thresh_individual_high): 
                            fn_y.append(2) 
                        else: 
                            fn_y.append(1) 
                    #is this a false positive? 
                    elif truth_sample and (maf > Params.maf_thresh_individual_low) and (maf_dict[snp][truth_sample][1] > Params.min_calls) and (maf_dict[snp][truth_sample][0] < Params.maf_thresh_individual_low): 
                        fp_x.append(snp_to_pos[snp]) 
                        if maf>(Params.maf_thresh_individual_high): 
                            fp_y.append(2) 
                        else: 
                            fp_y.append(1) 
                    #are there 2 minor alleles? 
                    elif maf > (Params.maf_thresh_individual_high):
                        xvals.append(snp_to_pos[snp]) 
                        yvals.append(2)
                    elif maf > (Params.maf_thresh_individual_low): 
                        xvals.append(snp_to_pos[snp]) 
                        yvals.append(1) 
        if len(xvals) >0:
            markerline, stemlines, baseline = stem(xvals, yvals, '-')
            plots.append((markerline, stemlines, baseline))
            setp(stemlines, 'color', 'b','linewidth',1.5)
            setp(markerline,'marker','None') 
        #plot proxy artist
        h, = plt.plot(1,1,color='b')
        proxies.append(h)
        if len(nc_x) > 0:
            #plt.scatter(nc_x,nc_y,s=30,facecolors="none",edgecolors='k')
            #plt.scatter(nc_x,nc_y,'k|',label='No Call')
            markerline_nc,stemlines_nc,baseline_nc=stem(nc_x,nc_y,'k|')#,label='No Call') 
            setp(stemlines_nc,'color','w','linewidth',0) 
            setp(markerline_nc,'marker','|','color','k','markersize',50,'markerfacecolor','k')
        h,=plt.plot(1,1,color='k',linestyle='o') 
        proxies.append(h)
        if len(fp_x) > 0:
            markerline_fp,stemlines_fp,baseline_fp=stem(fp_x,fp_y,'-')#,label='False Positive') 
            setp(stemlines_fp,'color','r','linewidth',1.5) 
            setp(markerline_fp,'marker','None','color','r') 
        h,=plt.plot(1,1,color='r') 
        proxies.append(h) 

        if len(fn_x) > 0:
            markerline_fn,stemlines_fn,baseline_fn=stem(fn_x,fn_y,'-.')#,label='False Negative') 
            setp(stemlines_fn,'color','r','linewidth',1.5) 
            setp(markerline_fn,'marker','None','color','r') 
        h,=plt.plot(1,1,color='r',linestyle='-.') 
        proxies.append(h) 

        if len(nop_x)>0: 
            markerline_nop,stemlines_nop,baseline_nop=stem(nop_x,nop_y,'-')
            setp(stemlines_nop,'color','k','linewidth',1.5) 
            setp(markerline_nop,'marker','None','color','k') 
        h,=plt.plot(1,1,color='k',linestyle='-.') 
        proxies.append(h) 

        plt.xlabel('SNP') 
        plt.ylabel('Minor Alleles') 
        plt.title(str(metadata[replicates[i]][0])+" "+str(metadata[replicates[i]][1]))

    # hide proxies    
    #plt.legend(proxies, legend_names,bbox_to_anchor=(0.,-0.3, 1.0,0.102),loc=3,ncol=4,mode="expand",borderaxespad=0., numpoints=1)
    #plt.legend(proxies, legend_names,bbox_to_anchor=(0.,-1, 1.0,0.102),loc=3,ncol=4,mode="expand",borderaxespad=0., numpoints=1)


    for h in proxies:
        h.set_visible(False)
    #plt.legend()#loc='lower center',bbox_to_anchor=(1,0.5))
    plt.subplots_adjust(hspace=0.8) 


    matplotlib.rcParams.update({'font.size': 22})


    f=plt.gcf() 
    f.set_size_inches(20,16) 
    plt.savefig(Params.output_details_plot+title,dpi=80)
    #param_hash=str(Params.maf_thresh_individual_low_id)+str(Params.maf_thresh_individual_amb_id)+str(Params.min_calls_id)+str(Params.ambiguous_thresh_id)+str(Params.strand_bias_thresh_id) 
    #internal_hash=hash(param_hash+internal_hash) 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+"',"+str(internal_hash)+",'"+str(description)+"','Replicate Details','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    print str(query) 
    args=open(Params.output_details_plot+title,'r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash='"+str(internal_hash)+"';"
    cur.execute(query) 
    details_image_id=cur.fetchone()[0]  
    return details_image_id


def outputHistImage(maf_dict,ma_snps,replicates,truth_sample,internal_hash,metadata,cur,user_id,folder_id,quality,locus_group): 
    hist_image_id=0

     #get file_name that the user will see 
    panels=set([]) 
    sources=set([]) 
    title="ReplicateHist"
    if truth_sample: 
        title=title+"_Reference_"+metadata[truth_sample][0]+"_"+metadata[truth_sample][1] 
        sources=sources.union(metadata[truth_sample][-1]) 
        panels.add(metadata[truth_sample][-2]) 
    for i in range(len(replicates)): 
        if replicates[i]!=truth_sample: 
            title=title+"_"+(metadata[replicates[i]])[0]+"_"+(metadata[replicates[i]])[1]
            sources=sources.union(metadata[replicates[i]][-1]) 
            panels.add(metadata[replicates[i]][-2])
    if len(title) > 250: 
        title="ReplicateHist_"+str(len(replicates))
        for panel in panels: 
            title=title+"_"+str(panel) 
        for source in sources: 
            title=title+"_"+str(source) 
    title=title+".png"

    #Populate the contributors and description fields 
    description="Quality:"+str(quality)+"\n"+"Locus Group:"+str(locus_group)+"\n"  
    if truth_sample: 
        description=description+"Reference:\n"+metadata[truth_sample][3]+" ("+metadata[truth_sample][0]+")"+ " "+metadata[truth_sample][1] +" ("+metadata[truth_sample][2]+")" +"\n" 
    description=description+"Replicates:"
    for rep in replicates: 
        if rep==truth_sample: 
            continue 
        description=description+"\n"+metadata[rep][3]+" (" + metadata[rep][0]+")"+ " "+metadata[rep][1]+" ("+metadata[rep][2]+")"


    numreps=len(replicates) 
    plt.clf() 
    for i in range(len(replicates)):
        plt.subplot(numreps,1,i+1) 
        ref=replicates[i]
        ref_name=str(metadata[ref][0])+" " + str(metadata[ref][1])
        comparison_names=[] 
        comp_dict=dict() 
        for i in range(3): 
            for j in range(3): 
                comp_dict[str(i)+str(j)]=[] 
        #get the names of the replicates that will be compared to the reference. 
        for j in range(len(replicates)): 
            comparison=replicates[j] 
            if ref==comparison:
                continue
            comp_name=str(metadata[comparison][0])+"_"+str(metadata[comparison][1])
            comparison_names.append(comp_name) 
            for k in range(3): 
                for m in range(3): 
                    comp_dict[str(k)+str(m)].append(0) 
            
            for snp in ma_snps: 
                ref_call=0 
                comp_call=0 
                if ref not in maf_dict[snp]: 
                    continue 
                if comparison not in maf_dict[snp]: 
                    continue 
                ref_maf,ref_count,ref_count_forward=maf_dict[snp][ref] 
                comp_maf,comp_count,comp_count_forward=maf_dict[snp][comparison] 
                if ref_count > Params.min_calls: 
                    if ref_maf > (Params.maf_thresh_individual_high): 
                        ref_call=2 
                    elif ref_maf > Params.maf_thresh_individual_low: 
                        ref_call=1 
                if comp_count > Params.min_calls: 
                    if comp_maf > (Params.maf_thresh_individual_high): 
                        comp_call=2 
                    elif comp_maf > Params.maf_thresh_individual_low: 
                        comp_call=1 
                comp_dict[str(ref_call)+str(comp_call)][-1]+=1 
        for key in comp_dict: 
            for v in range(len(comp_dict[key])):
                if len(ma_snps)==0: 
                    comp_dict[key][v]=0 
                else:
                    comp_dict[key][v]=float(comp_dict[key][v])/len(ma_snps) 
        indices=np.arange(0,len(comparison_names)*7,7)
        bar_width=1
        zero_one=plt.bar(indices,comp_dict['01'],bar_width,color='r',label='MM_Mm')
        zero_two=plt.bar(indices+bar_width,comp_dict['02'],bar_width,color='#FF7F00',label='MM_mm') 
        one_zero=plt.bar(indices+2*bar_width,comp_dict['10'],bar_width,color='#FFFF00',label='Mm_MM')
        #one_one=plt.bar(indices+3*bar_width,comp_dict['11'],bar_width,color='#00FF00',label='Mm_Mm') 
        one_two=plt.bar(indices+4*bar_width,comp_dict['12'],bar_width,color='#00FF00',label='Mm_mm')
        two_zero=plt.bar(indices+5*bar_width,comp_dict['20'],bar_width,color='#0000FF',label='mm_MM')
        two_one=plt.bar(indices+6*bar_width,comp_dict['21'],bar_width,color='#4B0082',label='mm_Mm')
        #two_two=plt.bar(indices+7*bar_width,comp_dict['22'],bar_width,color='k',label='mm_mm')
        

        plt.ylabel(ref_name.replace(' ','\n') )
        #plt.xlabel('Replicates') 
        plt.xticks([i for i in indices],comparison_names)
        plt.grid(True,lw=1,c='0.5') 
    plt.legend(bbox_to_anchor=(1,2.5),loc='center left',ncol=1,mode="expand",borderaxespad=0)
    #plt.legend(loc='center left',bbox_to_anchor=(1,0.5)) 
    plt.subplots_adjust(left=0.07,right=0.8) 
    f=plt.gcf() 
    f.set_size_inches(12,8) 
    plt.savefig(Params.output_histogram+title,dpi=80)
    #param_hash=str(Params.maf_thresh_individual_low_id)+str(Params.maf_thresh_individual_amb_id)+str(Params.min_calls_id)+str(Params.ambiguous_thresh_id)+str(Params.strand_bias_thresh_id) 
    #internal_hash=hash(param_hash+internal_hash) 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+"',"+str(internal_hash)+",'"+str(description)+"','Replicate Histogram','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=open(Params.output_histogram+title,'r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash='"+str(internal_hash)+"';"
    cur.execute(query) 
    hist_image_id=cur.fetchone()[0]  
    return hist_image_id 
