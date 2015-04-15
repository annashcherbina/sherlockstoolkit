import operator; 
import Params
from openpyxl import Workbook,load_workbook 

#Generate an output Excel file for ancestry calculation 
def createOutputFile(aggregateAncestries,ethnicity_id_to_name,regions,cur,user_id,folder_id,parameter_group,quality,locus_group): 
    print "generating output\n" 
    #individual metadata for the output file 
    samples=list(aggregateAncestries.keys()) 
    samples.sort() 
    person_samples=[] 
    people=[] 
    sources=set([])  
    panels=set([])  
    barcodes=[] 
    exp_name=[] 
    exp_hash=[] 
    
    wb=Workbook()
    ws=wb.get_active_sheet() 
    ws.title="Population_Significant"
    c=ws.cell(row=0,column=0) 
    c.value="Subject" 
    c=ws.cell(row=0,column=1) 
    c.value="Barcode" 
    c=ws.cell(row=0,column=2) 
    c.value="Ancestry" 
    c=ws.cell(row=0,column=3) 
    c.value="Percent"

    ws2=wb.create_sheet() 
    ws2.title="Region_Significant"
    c=ws2.cell(row=0,column=0) 
    c.value="Subject" 
    c=ws2.cell(row=0,column=1) 
    c.value="Barcode" 
    c=ws2.cell(row=0,column=2) 
    c.value="AncestryRegion" 
    c=ws2.cell(row=0,column=3) 
    c.value="Percent"


    ws3=wb.create_sheet() 
    ws3.title="Population_Details"
    c=ws3.cell(row=0,column=0) 
    c.value="Subject" 
    c=ws3.cell(row=0,column=1) 
    c.value="Barcode" 
    c=ws3.cell(row=0,column=2) 
    c.value="AncestryRegion" 
    c=ws3.cell(row=0,column=3) 
    c.value="Percent"

    ws4=wb.create_sheet() 
    ws4.title="Region_Details"
    c=ws4.cell(row=0,column=0) 
    c.value="Subject" 
    c=ws4.cell(row=0,column=1) 
    c.value="Barcode" 
    c=ws4.cell(row=0,column=2) 
    c.value="AncestryRegion" 
    c=ws4.cell(row=0,column=3) 
    c.value="Percent"
    r=0
    i=0
    for sample in samples: 
        r+=1
        query="select id,person_id from person_samples where sample_id="+str(sample)+";" 
        #print str(query) 
        cur.execute(query) 
        [person_sample_id,person_id]=cur.fetchone()
        person_samples.append(person_sample_id) 
        query="select id_code,source from people where id="+str(person_id)+";" 
        #print str(query) 
        cur.execute(query) 
        person_details=cur.fetchone() 
        person=person_details[0] 
        source=person_details[1] 
        sources.add(source) 
        people.append(person) 
        query="select barcode_id from samples where id="+str(sample)+";" 
        cur.execute(query) 
        barcode_id=cur.fetchone()[0]
        query="select name from barcodes where id="+str(barcode_id)+";" 
        cur.execute(query) 
        barcode=cur.fetchone()[0] 
        barcodes.append(barcode) 
        query="select experiment_id from samples where id="+str(sample)+";" 
        cur.execute(query)
        experiment_id=cur.fetchone()[0]
        query="select name,hash_name,primer_panel_id from experiments where id="+str(experiment_id)+";" 
        cur.execute(query) 
        experiment_name,experiment_hash,panel_id=cur.fetchone()
        exp_name.append(experiment_name) 
        exp_hash.append(experiment_hash) 
        query="select name from panels where id="+str(panel_id)+";" 
        cur.execute(query) 
        panel_name=cur.fetchone()[0] 
        panels.add(panel_name) 
        probSum=0

        #keep track of unique regional ancestries: 
        regional_ancestries=dict()
        finalSum=0 
        print "aggregateAncestries:"+str(aggregateAncestries[sample]) 
        for ancestor in aggregateAncestries[sample]: 
            aggregateAncestries[sample][ancestor]=aggregateAncestries[sample][ancestor]/Params.runxtimes; 
            finalSum+=aggregateAncestries[sample][ancestor] 
            regional_ancestry=regions[ancestor] 
            if regional_ancestry not in regional_ancestries: 
                regional_ancestries[regional_ancestry]=float(aggregateAncestries[sample][ancestor])
            else: 
                regional_ancestries[regional_ancestry]+=float(aggregateAncestries[sample][ancestor]) 
        for region in regional_ancestries: 
            regional_ancestries[region]=regional_ancestries[region]/float(finalSum) 
        for ancestor in aggregateAncestries[sample]: 
            aggregateAncestries[sample][ancestor]=aggregateAncestries[sample][ancestor]/float(finalSum)
      
        #reverse sort by freqs 
        ancestries_sorted = sorted(aggregateAncestries[sample].iteritems(), key=operator.itemgetter(1))
        ancestries_sorted.reverse(); 
        regions_sorted=sorted(regional_ancestries.iteritems(),key=operator.itemgetter(1)) 
        regions_sorted.reverse() 

        #check to see if an entry for this subject already exists. If not, create a new entry. Otherwise,do nothing. 
        query="select * from ancestries where person_sample_id="+str(person_sample_id)+";"
        cur.execute(query) 
        existing=cur.fetchall()
        if len(existing)==0: 
            for entry in ancestries_sorted: 
                ancestry_id=entry[0] 
                percent=round(100*entry[1],2) 
                geo_id=regions[ancestry_id] 
                source_description=exp_name[i]+ " ("+str(exp_hash[i])+") "+str(barcodes[i])+" ("+str(people[i])+")"
                #print "source_description:"+str(source_description) 
                query="insert into ancestries (person_sample_id,ethnicity_id,geographic_id,percent,source_description) values("+str(person_sample_id)+","+str(ancestry_id)+","+str(geo_id)+","+str(percent)+",'"+str(source_description)+"');"
                print str(query) 
                cur.execute(query)
                
        i+=1
        other=0      
        for ancestor in aggregateAncestries[sample]:
            if aggregateAncestries[sample][ancestor] < Params.toolow: 
                other+=aggregateAncestries[sample][ancestor] 
                aggregateAncestries[sample][ancestor]=0; 
        other_regions=0 
        for region in regional_ancestries: 
            if regional_ancestries[region]< Params.toolow: 
                other_regions+=regional_ancestries[region] 
                regional_ancestries[region]=0 

        reduced_ancestries = dict(); 
        reduced_ancestries['other']=other 
        for ancestor in aggregateAncestries[sample]: 
            if aggregateAncestries[sample][ancestor] > 0: 
                reduced_ancestries[ancestor]=aggregateAncestries[sample][ancestor]; 
        #reverse sort by freqs 
        sorted_reduced_ancestries = sorted(reduced_ancestries.iteritems(), key=operator.itemgetter(1))
        sorted_reduced_ancestries.reverse(); 

        reduced_regions=dict(); 
        reduced_regions["other"]=other_regions; 
        for region in regional_ancestries: 
            if regional_ancestries[region]> 0: 
                reduced_regions[region]=regional_ancestries[region] 

        sorted_reduced_regions=sorted(reduced_regions.iteritems(),key=operator.itemgetter(1)) 
        sorted_reduced_regions.reverse() 
    
                
        #write the data to the Excel file         
        for sheetname in [ws,ws2,ws3,ws4]: 
            c=sheetname.cell(row=r,column=0)
            c.value=person
            c=sheetname.cell(row=r,column=1) 
            c.value=barcode; 


        for colname in range(len(ancestries_sorted)):
            entry=ancestries_sorted[colname] 
            query="select name from ethnicities where id="+str(entry[0])+";" 
            cur.execute(query) 
            ancestry_name=cur.fetchone()[0] 
            c=ws3.cell(row=r,column=2*colname+2)
            c.value=ancestry_name 
            c=ws3.cell(row=r,column=2*colname+3)
            c.value=round(100*entry[1],2)
        for colname in range(len(sorted_reduced_ancestries)): 
            entry=sorted_reduced_ancestries[colname] 
            if entry[0]=="other": 
                ancestry_name="other" 
            else: 
                query="select name from ethnicities where id="+str(entry[0])+";" 
                cur.execute(query) 
                ancestry_name=cur.fetchone()[0] 
            c=ws.cell(row=r,column=2*colname+2) 
            c.value=ancestry_name 
            c=ws.cell(row=r,column=2*colname+3) 
            c.value=round(100*entry[1],2)


        for colname in range(len(regions_sorted)): 
            entry=regions_sorted[colname] 
            query="select region_name from geographics where id="+str(entry[0])+";" 
            cur.execute(query) 
            print str(query) 
            region_name=cur.fetchone()[0] 
            c=ws4.cell(row=r,column=2*colname+2) 
            c.value=region_name 
            c=ws4.cell(row=r,column=2*colname+3) 
            c.value=round(100*entry[1],2) 

        for colname in range(len(sorted_reduced_regions)): 
            entry=sorted_reduced_regions[colname] 
            if entry[0]=="other": 
                region_name="other" 
            else: 
                query="select region_name from geographics where id="+str(entry[0])+";" 
                cur.execute(query) 
                region_name=cur.fetchone()[0] 
            c=ws2.cell(row=r,column=2*colname+2) 
            c.value=region_name
            c=ws2.cell(row=r,column=2*colname+3) 
            c.value=round(100*entry[1],2)

    ws5=wb.create_sheet()
    query="select name from users where id="+str(user_id)+";" 
    cur.execute(query) 
    username=cur.fetchone() 
    if username!=None: 
        username=username[0] 
    ws5.title="Parameters" 
    c=ws5.cell(row=0,column=0) 
    c.value=str(parameter_group)+","+str(username) 
    c=ws5.cell(row=1,column=0) 
    c.value="Minimum Reads Per Locus" 
    c=ws5.cell(row=1,column=1) 
    c.value=Params.toolowthreshold
    
    c=ws5.cell(row=2,column=0) 
    c.value="Reference - MAF for 0 Minor Alleles (upper bound)"
    c=ws5.cell(row=2,column=1) 
    c.value=Params.maf_thresh_individual_low  


    c=ws5.cell(row=3,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (lower bound)"
    c=ws5.cell(row=3,column=1) 
    c.value=Params.maf_thresh_individual_amb

    c=ws5.cell(row=4,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (upper bound)"
    c=ws5.cell(row=4,column=1) 
    c.value=Params.maf_thresh_individual_amb_upper

    c=ws5.cell(row=5,column=0) 
    c.value="Reference - MAF for 2 Minor Alleles (lower bound)"
    c=ws5.cell(row=5,column=1) 
    c.value=Params.maf_thresh_individual_high

    c=ws5.cell(row=6,column=0) 
    c.value="Minimum number of Kidd SNPs present (out of 128)"
    c=ws5.cell(row=6,column=1) 
    c.value=Params.minsnpspresent


    c=ws5.cell(row=7,column=0) 
    c.value="Minimum ancestry contributions to report (percent)"
    c=ws5.cell(row=7,column=1) 
    c.value=Params.contributing_thresh

    c=ws5.cell(row=8,column=0) 
    c.value="Genetic algorithm runs for subject (these will be averaged)"
    c=ws5.cell(row=8,column=1) 
    c.value=Params.runxtimes

    c=ws5.cell(row=9,column=0) 
    c.value="Minimum ancestry contributions to report (number)"
    c=ws5.cell(row=9,column=1) 
    c.value=Params.maxtoreport
    
    c=ws5.cell(row=10,column=0) 
    c.value="Global Strand Bias Threshold"
    c=ws5.cell(row=10,column=1) 
    c.value=Params.strand_bias_thresh 

    c=ws5.cell(row=11,column=0) 
    c.value="Global Ambiguous Threshold" 
    c=ws5.cell(row=11,column=1) 
    c.value=Params.ambiguous_thresh 

    c=ws5.cell(row=12,column=0) 
    c.value="Quality Threshold" 
    c=ws5.cell(row=12,column=1) 
    c.value = quality 

    c=ws5.cell(row=13,column=0) 
    c.value="Locus Group" 
    c=ws5.cell(row=13,column=1)
    loc_group_name="Unspecified"
    if str(locus_group)=='0': 
        c.value=loc_group_name
    else: 
        query="Select name from loci_groups where id="+str(locus_group)+";" 
        cur.execute(query) 
        loc_group_name=cur.fetchone()[0] 
        c.value=loc_group_name 
        #create a tab listing all the loci in the locus group 
        ws6=wb.create_sheet() 
        ws6.title="Locus Group"
        c=ws.cell(row=0,column=0) 
        c.value=loc_group_name 
        query="select locus_id from loci_loci_groups where loci_group_id="+str(locus_group)+";"
        cur.execute(query) 
        loci_in_group=cur.fetchall()
        loci_string='|'.join([str(i[0]) for i in loci_in_group])
        query="select name from loci where id REGEXP '("+loci_string+")';" 
        cur.execute(query) 
        loci_names_in_group=cur.fetchall()
        for i in range(1,len(loci_names_in_group)-1): 
            c=ws.cell(row=i,column=0) 
            c.value=loci_names_in_group[i-1]


    title="Ancestries"
    param_hash=str(Params.toolowthreshold_id)+str(Params.contributing_thresh_id)+str(Params.runxtimes_id)+str(Params.maxtoreport_id)+str(Params.minsnpspresent_id)+str(Params.maf_thresh_individual_low_id)+str(Params.maf_thresh_individual_amb_id)+str(Params.maf_thresh_individual_amb_upper_id)+str(Params.maf_thresh_individual_high_id) +str(Params.strand_bias_thresh)+str(Params.ambiguous_thresh)+str(locus_group) 
    internal_hash=param_hash+"Ancestries" 
    description="Quality Cutoff:"+str(quality)+"\n"+"Locus Group:"+str(loc_group_name)+"\n"
    for i in range(len(samples)):
        title=title+"_"+str(exp_hash[i])+"_"+str(barcodes[i])
        internal_hash=internal_hash+"_"+str(person_samples[i]) 
        description=description+exp_name[i]+ " ("+str(exp_hash[i])+") "+str(barcodes[i])+" ("+str(people[i])+")\n" 
    if len(title)>250: 
        title="Ancestries_"+str(len(samples))
        for p in panels: 
            title=title+"_"+str(p) 
        for source in sources: 
            title=title+"_"+str(source) 
       
    title=title+".xlsx" 
    print "internal_hash:"+str(internal_hash) 
    wb.save(Params.output_excel+title)

    #Insert Excel file into the database.
    print "Internal hash inserted:"+str(internal_hash) 
    internal_hash=hash(internal_hash)
    print "inserting hash:"+str(internal_hash) 
    query="INSERT into attachments (user_id,folder_id,file_name,internal_hash,description,file_type,content_type,updated_at,codebase_version,contents) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+"',"+str(internal_hash)+",'"+str(description)+"','Ancestry','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    args=open(Params.output_excel+title,'r').read()  
    cur.execute(query,args)
    query="select id from attachments where internal_hash="+str(internal_hash)+";" 
    cur.execute(query) 
    attach_id=cur.fetchone()[0]
    #add the associated attachment id to the ancestries entries 
    #return the person-ids for the ancestry computation 
    return person_samples,attach_id 
