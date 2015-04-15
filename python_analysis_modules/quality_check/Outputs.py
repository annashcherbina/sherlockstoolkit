import Params; 
import matplotlib 
matplotlib.use('Agg')
import matplotlib.pyplot as plt 
import matplotlib.colors as colors
import matplotlib.cm as cmx
import matplotlib
import numpy as np; 
from openpyxl import Workbook,load_workbook 

#finds the title, description, and internal name associated with the sample image. 
def get_metadata(samples,cur,quality,locus_group_name,locus_group):
    samples.sort() 
    sample_metadata=dict() 
    description="Quality:"+str(quality)+"\n"+ "Locus Group:"+str(locus_group_name)+"\n" 
    param_hash=str(Params.min_calls_id) +str(locus_group) 
    internal_hash=param_hash+"qc" 
    title="QualityCheck"
    panels=set([]) 
    sources=set([]) 
    for sample in samples: 
        query="select experiment_id,barcode_id from samples where id="+str(sample)+";" 
        cur.execute(query) 
        exp_dets=cur.fetchone()
        exp_id=exp_dets[0] 
        barcode_id=exp_dets[1] 
        query="select name from barcodes where id="+str(barcode_id)+";" 
        cur.execute(query) 
        barcode_name=cur.fetchone()[0] 
        query="select name,hash_name,primer_panel_id from experiments where id="+str(exp_id)+";" 
        cur.execute(query) 
        experiment_name,hash_name,primer_panel_id=cur.fetchone()
        query="select person_id from person_samples where sample_id="+str(sample)+";"; 
        cur.execute(query) 
        person_id=cur.fetchall() 
        if len(person_id)>1: 
            person_name="Mixture" 
        else: 
            person_id=person_id[0][0] 
            query="select id_code,source from people where id="+str(person_id)+";" 
            #print str(query) 
            cur.execute(query) 
            person_details=cur.fetchone()
            person_name=person_details[0] 
            source=person_details[1] 
            sources.add(source) 
        query="select name from panels where id="+str(primer_panel_id)+";" 
        cur.execute(query) 
        panel_id=cur.fetchone()[0].split('_')[0] 
        panels.add(panel_id) 
        internal_hash=internal_hash+"_"+str(sample) 
        title=title+"_"+hash_name+"_"+barcode_name
        description=description+experiment_name+" ("+hash_name+") "+barcode_name+" ("+person_name+")"+"\n" 
        sample_metadata[sample]=[hash_name,experiment_name,barcode_name,person_name,panel_id] 
    if len(title)>250: 
        title="QualityCheck_"+str(len(samples))
        for p in panels: 
            title=title+"_"+str(p) 
        for source in sources: 
            title=title+"_"+str(source) 
        
    return sample_metadata ,internal_hash,title,description

def generate_attachment(sample_to_ma,cur,user_id,folder_id,parameter_group,quality,locus_group_name,locus_group_snps,locus_group,snp_to_freq): 
    sample_metadata,internal_hash,title,description=get_metadata(sample_to_ma.keys(),cur,quality,locus_group_name,locus_group)
    wb=Workbook() 
    ws=wb.get_active_sheet() 
    ws.title="MAF Sorted" 
    samples=sample_to_ma.keys() 
    for i in range(len(samples)): 
        c=ws.cell(row=0,column=i) 
        c.value=sample_metadata[samples[i]][-2]  
    for i in range(len(samples)):
        for j in range(len(sample_to_ma[samples[i]])): 
            c=ws.cell(row=j+1,column=i) 
            c.value=sample_to_ma[samples[i]][j] 
    ws2=wb.create_sheet() 
    query="select name from users where id="+str(user_id)+";" 
    cur.execute(query) 
    username=cur.fetchone() 
    if username!=None: 
        username=username[0] 
    ws2.title="Parameters" 
    c=ws2.cell(row=0,column=0) 
    c.value=parameter_group+","+str(username) 
    c=ws2.cell(row=1,column=0) 
    c.value="Minimum Reads Per Locus" 
    c=ws2.cell(row=1,column=1) 
    c.value=Params.min_calls
    c=ws2.cell(row=2,column=0) 
    c.value="Quality" 
    c=ws2.cell(row=2,column=1) 
    c.value=quality 
    c=ws2.cell(row=3,column=0) 
    c.value="Locus Group" 
    c=ws2.cell(row=3,column=1) 
    c.value=locus_group_name 
    if len(locus_group_snps) > 0: 
        ws3=wb.create_sheet() 
        ws3.title="Locus Group" 
        c=ws3.cell(row=0,column=0) 
        c.value=locus_group_name 
        snp_string='|'.join([str(i) for i in locus_group_snps])
        query="select name from loci where id REGEXP '("+snp_string+")';"
        cur.execute(query) 
        snp_names=[i[0] for i in cur.fetchall()]        
        for i in range(len(snp_names)): 
            c=ws3.cell(row=i+1,column=0) 
            c.value = snp_names[i] 
    ws3=wb.create_sheet()
    ws3.title="MAF by SNP"
    snp_ids=snp_to_freq.keys() 
    snp_string="|".join([str(i) for i in snp_ids])
    #print "snp_string:"+str(snp_string) 
    query="select id,name from loci where id REGEXP '("+snp_string+")';"
    cur.execute(query) 
    idmap=cur.fetchall() 
    #print "snp_ids:"+str(snp_ids) 
    #print "idmap:"+str(idmap) 

    
    c=ws3.cell(row=0,column=0) 
    c.value="SNP"
    for i in range(len(samples)): 
        c=ws3.cell(row=0,column=i*3+1) 
        c.value="MAF "+str(sample_metadata[samples[i]][-2]) 
        c=ws3.cell(row=0,column=i*3+2) 
        c.value="READS "+str(sample_metadata[samples[i]][-2])
        c=ws3.cell(row=0,column=i*3+3) 
        c.value="FORWARD READS"+str(sample_metadata[samples[i]][-2])
    for j in range(len(idmap)): 
        rowval=j+1 
        snp_id=idmap[j][0]
        if snp_id not in snp_to_freq: 
            continue 
        #print "snp_id:"+str(snp_id) 
        c=ws3.cell(row=rowval,column=0) 
        c.value=idmap[j][1] 
        for i in range(len(samples)): 
            sample_name=samples[i]
            c=ws3.cell(row=rowval,column=i*3+1) 
            #print str (snp_to_freq[snp_id]) 
            c.value=snp_to_freq[snp_id][sample_name][0]
            c=ws3.cell(row=rowval,column=i*3+2) 
            c.value=snp_to_freq[snp_id][sample_name][1]
            c=ws3.cell(row=rowval,column=i*3+3) 
            c.value=snp_to_freq[snp_id][sample_name][2] 
    wb.save("output/"+title+".xlsx")
    internal_hash=hash(internal_hash) 
    query="INSERT into attachments (user_id,folder_id,file_name,internal_hash,description,file_type,content_type,updated_at,codebase_version,contents) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+".xlsx"+"',"+str(internal_hash)+",'"+str(description)+"','Quality Check','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    args=(open("output/"+title+".xlsx",'r').read(),) 
    cur.execute(query,args)
    query="select id from attachments where internal_hash='"+str(internal_hash)+"';"
    cur.execute(query) 
    attachment_id=cur.fetchone()[0] 
    return attachment_id 

def generate_image(sample_to_ma,cur,user_id,folder_id,quality,locus_group_name,locus_group): 
    im_id=None 

    jet = cm =plt.get_cmap('jet') 
    cNorm  = colors.Normalize(vmin=0, vmax=len(sample_to_ma)-1)
    scalarMap = cmx.ScalarMappable(norm=cNorm, cmap=jet)
    
    fig = plt.figure()
    ax1 = fig.add_subplot(111)
    sample_metadata,internal_hash,title,description=get_metadata(sample_to_ma.keys(),cur,quality,locus_group_name,locus_group)
    title="output/"+title+".png" 
    counter=0
    keys=sample_to_ma.keys() 
    for idx in range(len(sample_to_ma)):
        sample=keys[idx] 
        colorVal = scalarMap.to_rgba(idx)
        xvals=range(1,len(sample_to_ma[sample])+1) 
        yvals=sample_to_ma[sample]
        label_name=sample_metadata[sample][0]+ " "+sample_metadata[sample][2]+" " + str(sample_metadata[sample][3]) 
        ax1.scatter(xvals,yvals,s=5,color=colorVal,label=label_name)
        counter+=1
    plt.legend(bbox_to_anchor=(0.95,1),loc=2,borderaxespad=0)#loc='upper left');
    plt.xlabel('SNP') 
    plt.ylabel('Minor Allele Frequency')
    plt.title('Minor Allele Frequency Check')
    matplotlib.rcParams.update({'font.size':22}) 

    f=plt.gcf() 
    f.set_size_inches(20,8) 
    plt.savefig(title,dpi=80)
    internal_hash=hash(internal_hash) 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title).replace("output/","")+"',"+str(internal_hash)+",'"+str(description)+"','Quality Check','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=open(title,'r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash="+str(internal_hash)+";"
    cur.execute(query) 
    im_id=cur.fetchone()[0]  
    return im_id
