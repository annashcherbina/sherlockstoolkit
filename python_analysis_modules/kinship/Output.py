import Params; 
import matplotlib 
matplotlib.use('Agg')
import matplotlib.pyplot as plt 
from openpyxl import Workbook,load_workbook
import pandas as pd; 
import time
import sys, os
from pylab import *  # for plotting
from numpy.random import *  # for random sampling
import cPickle

# We need to import the graph_tool module
import graph_tool
import graph_tool.stats
import graph_tool.draw
from pedigree_graph2 import * 
import pygraphviz as PG 




def get_metadata(samples,cur,training_module,quality,locus_group): 
    samples.sort() 
    sample_metadata=dict() 
    description="Training Module:"+str(training_module)+"\n"+"Quality Cutoff:"+str(quality)+"\n"+ "Locus Group:"+str(locus_group)+'\n' 
    internal_hash="kin"
    param_hash=str(Params.max_0_ma_id)+str(Params.min_1_ma_id)+str(Params.max_1_ma_id)+str(Params.min_2_ma_id)+str(Params.min_calls_id)+str(Params.ambiguous_thresh_id)+str(Params.strand_bias_thresh_id)+str(Params.max_impossible_id)+str(locus_group)+str(training_module)+str(time.strftime("%H:%M:%S"))
    internal_hash=internal_hash+param_hash 
    title="Kinship"
    sources=set([]) 
    panels=set([]) 
    for sample in samples: 
        query="select experiment_id,barcode_id from samples where id="+str(sample)+";" 
        cur.execute(query) 
        exp_dets=cur.fetchone()
        exp_id=exp_dets[0] 
        barcode_id=exp_dets[1] 
        query="select name from barcodes where id="+str(barcode_id)+";" 
        cur.execute(query) 
        barcode_name=cur.fetchone()[0] 
        query="select name,hash_name,primer_panel_id from experiments where id="+str(exp_id)+";" 
        cur.execute(query) 
        experiment_name,hash_name,primer_panel_id=cur.fetchone()
        query="select person_id from person_samples where sample_id="+str(sample)+";"; 
        cur.execute(query) 
        person_id=cur.fetchall() 
        person_id=person_id[0][0] 
        query="select id_code,source from people where id="+str(person_id)+";" 
        print str(query) 
        cur.execute(query) 
        person_details=cur.fetchone()
        person_name=person_details[0] 
        person_source=person_details[1]
        sources.add(person_source) 
        query="select name from panels where id="+str(primer_panel_id)+";" 
        cur.execute(query) 
        panel_id=cur.fetchone()[0].split('_')[0] 
        panels.add(panel_id) 
        internal_hash=internal_hash+"_"+str(sample) 
        title=title+"_"+hash_name+"_"+barcode_name
        description=description+experiment_name+" ("+hash_name+") "+barcode_name+" ("+person_name+")"+"\n" 
        sample_metadata[sample]=[hash_name,experiment_name,barcode_name,person_name,panel_id] 
    if len(title)>=250: 
        num_samples=len(samples)
        title="Kinship_"+str(num_samples)
        for p in panels: 
            title=title+"_"+str(p) 
        for source in sources: 
            title=title+"_"+str(source) 
        
    return sample_metadata ,internal_hash,title,description 

def generatePedigreeGraph(rel_dict,cur,samples,user_id,folder_id,training_module,quality,locus_group_name,allele_combo_dict,gender_dict,ancestry_dict): 
    sample_metadata,internal_hash,title,description=get_metadata(samples,cur,training_module,quality,locus_group_name)    
    p=PedigreeGraph(rel_dict,allele_combo_dict,gender_dict,ancestry_dict) 
    p.create_vertices() 
    p.add_edges() 
    nuclear_families=p.impute_direction() 
    #generate a newick string for pedigree visualization 
    tree=p.build_tree(nuclear_families) 
    #connect nuclear families based on 2nd degree relationships 
    tree=p.connect_nuclear(tree) 
    p.draw_graph() 
    A=PG.AGraph(directed=True,strict=False) 
    for node in tree: 
        for child in tree[node].children: 
            A.add_edge(node,child) 
            n=A.get_node(child) 
            if child not in p.gender_dict: 
                n.attr['shape']='diamond'
                #if the child of this node has another parent with a known gender, we can infer the gender 
                temp_children=tree[child].children 
                for c_temp in temp_children: 
                    for p_temp in tree[c_temp].parents: 
                        if p_temp in p.gender_dict: 
                            if p.gender_dict[p_temp]=='m': 
                                n.attr['shape']='circle'
                                break 
                            elif p.gender_dict[p_temp]=='f': 
                                n.attr['shape']='square' 
                                break 
            elif p.gender_dict[child]=='m': 
                n.attr['shape']='square'
            else: 
                n.attr['shape']='circle' 
            n=A.get_node(node) 
            if node not in p.gender_dict: 
                n.attr['shape']='diamond'
                #if the child of this node has another parent with a known gender, we can infer the gender 
                temp_children=tree[node].children 
                for c_temp in temp_children: 
                    for p_temp in tree[c_temp].parents: 
                        if p_temp in p.gender_dict: 
                            if p.gender_dict[p_temp]=='m': 
                                n.attr['shape']='circle' 
                            elif p.gender_dict[p_temp]=='f': 
                                n.attr['shape']='square' 
                                break 

            elif p.gender_dict[node]=='m': 
                n.attr['shape']='square'
            else: 
                n.attr['shape']='circle' 
    #randomly assign gender in cases where both parent nodes are unknown 
    for node in tree: 
        unknown_parents=[] 
        for p in tree[node].parents: 
            n=A.get_node(p) 
            if n.attr['shape']=='diamond': 
                unknown_parents.append(p) 
        if len(unknown_parents)>1: 
            p1=unknown_parents[0] 
            n=A.get_node(p1) 
            n.attr['shape']='square' 
            p2=unknown_parents[1] 
            n=A.get_node(p2) 
            n.attr['shape']='circle' 

    A.write('pedigree.dot') 
    A.layout(prog='dot') 
    A.draw('pedigree.png') 

    
    internal_hash=hash(internal_hash) 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title).replace("output/","")+".png"+"',"+str(internal_hash)+",'"+str(description)+"','Pedigree Graph','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=open('pedigree_graph.png','r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash="+str(internal_hash)+" and image_type like \"%Pedigree Graph%\";"
    cur.execute(query) 
    im_id=cur.fetchone()[0]  

    #insert the computed pedigree into the database 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title).replace("output/","")+".png"+"',"+str(internal_hash)+",'"+str(description)+"','Family Tree','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=open('pedigree.png','r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash="+str(internal_hash)+" and image_type like \"%Family Tree%\";"
    cur.execute(query) 
    im_id2=cur.fetchone()[0]  

    return im_id, im_id2

    

def plotKinCoef(pair_dict,cur,samples,user_id,folder_id,training_module,quality,locus_group_name):
    sample_metadata,internal_hash,title,description=get_metadata(samples,cur,training_module,quality,locus_group_name) 
    fig=plt.figure() 
    ax1=fig.add_subplot(111) 
    indexvals=range(int((len(samples)-1)*len(samples)*.5))
    for i in indexvals:
        pScore=pair_dict['IBS0'][i]
        tScore=pair_dict['KinC'][i]
        pmark=pair_dict['marker'][i] 
        ax1.plot(pScore,tScore,pmark,markersize=12) 

    ax1.plot([0,1],[0,0],'--k')
    ax1.plot([0,1],[.0625,.0625],'--k')
    ax1.plot([0,1],[.125, .125],'--k')
    ax1.plot([0,1],[.25,.25],'--k')
    ax1.plot([0,1],[.5,.5],'--k')

    ax1.plot([.25,.25],[0,.5],'--k')
    ax1.plot([.5,.5],[0,.5],'--k')
    ax1.plot([.75,.75],[0,.5],'--k')
    ax1.plot([1,1],[0,.5],'--k')
    ax1.set_xlim([0,1.2])
    circle_parents=plt.Circle((0.05,0.25),0.05,lw=5,ls='dashed',color='g',fill=False) 
    fig.gca().add_artist(circle_parents) 
    circle_siblings=plt.Circle((0.25,0.25),0.05,lw=5,ls='dashed',color='k',fill=False) 
    fig.gca().add_artist(circle_siblings)
    circle_degree2=plt.Circle((0.5,0.125),0.05,lw=5,ls='dashed',color='m',fill=False) 
    fig.gca().add_artist(circle_degree2)
    circle_degree3=plt.Circle((0.75,0.0625),0.05,lw=5,ls='dashed',color='b',fill=False) 
    fig.gca().add_artist(circle_degree3)
    circle_unrelated=plt.Circle((1,0),0.12,lw=5,ls='dashed',color='r',fill=False) 
    fig.gca().add_artist(circle_unrelated)
    
    
    
    #one marker for each relationship, as defined in Params.py 
    L1 = matplotlib.lines.Line2D((0,1),(0,1),color='g',marker='s',markersize=10)
    L2 = matplotlib.lines.Line2D((0,1),(0,1),color='k',marker='o',markersize=10)
    L3=  matplotlib.lines.Line2D((0,1),(0,1),color='m',marker='o',markersize=10)
    L4 = matplotlib.lines.Line2D((0,1),(0,1),color='m',marker='^',markersize=10)
    L5 = matplotlib.lines.Line2D((0,1),(0,1),color='m',marker='d',markersize=10)
    L55 = matplotlib.lines.Line2D((0,1),(0,1),color='b',marker='s',markersize=10)
    L6 = matplotlib.lines.Line2D((0,1),(0,1),color='b',marker='^',markersize=10)
    L7 = matplotlib.lines.Line2D((0,1),(0,1),color='b',marker='d',markersize=10)
    L8 = matplotlib.lines.Line2D((0,1),(0,1),color='b',marker='>',markersize=10)
    L9 = matplotlib.lines.Line2D((0,1),(0,1),color='r',marker='x',markersize=10)
    L10 = matplotlib.lines.Line2D((0,1),(0,1),color='k',marker='x',markersize=10)
    L11 = matplotlib.lines.Line2D((0,1),(0,1),color='r',marker='^',markersize=10)
    L12 = matplotlib.lines.Line2D((0,1),(0,1),color='g',marker='x',markersize=10)
    L13 = matplotlib.lines.Line2D((0,1),(0,1),color='#20b2aa',marker='o',markersize=10)
    L14 = matplotlib.lines.Line2D((0,1),(0,1),color='g',marker='<',markersize=10) 
    L15 = matplotlib.lines.Line2D((0,1),(0,1),color='#99cc32',marker='x',markersize=10) 
    L16 = matplotlib.lines.Line2D((0,1),(0,1),color='#838b83',marker='*',markersize=10) 
    L17 = matplotlib.lines.Line2D((0,1),(0,1),color='#ffa07a',marker='*',markersize=10) 
    L18 = matplotlib.lines.Line2D((0,1),(0,1),color='b',marker='x',markersize=10) 
    L19 = matplotlib.lines.Line2D((0,1),(0,1),color='k',marker='*',markersize=10) 
    L20 = matplotlib.lines.Line2D((0,1),(0,1),color='r',marker='*',markersize=10) 

    #plt.legend((L1,L2,L3,L4,L5,L6,L7,L8,L9,L10,L11,L12,L13,L14,L15,L16,L17),('Parent/Child','Siblings','Half Siblings','Grandparent/child','Aunt/Uncle','Great Grandparent','Great Aunt/Uncle','Cousin','Unrelated','Not Calculated','4th Degree','5th Degree','6th Degree','7th Degree','8th Degree', '9th Degree','10th Degree'),bbox_to_anchor=(0.95, 1), loc=2, borderaxespad=0.)
    plt.legend((L1,L2,L3,L4,L5,L55,L6,L7,L8,L9,L11,L12,L13,L14,L15,L16,L17,L18,L19,L20),('Parent/Child','Siblings','Half Siblings','Grandparent/child','Aunt/Uncle','Half Aunt/Uncle','Great Grandparent','Great Aunt/Uncle','Cousin','Unrelated','4th Degree','5th Degree','6th Degree','7th Degree','8th Degree','9th Degree','>=10th Degree'),bbox_to_anchor=(0.95, 1), loc=2, borderaxespad=0.)
    plt.title('KING algorithm',size=20,weight='bold')
    plt.ylabel('Kinship Coef',size=20,weight='bold')
    plt.xlabel('Pr(Zero Identity by State)',size=20,weight='bold')
    matplotlib.rcParams.update({'font.size': 22})
    f=plt.gcf() 
    f.set_size_inches(30,10) 
    plt.savefig(Params.output_image+title+'.png',dpi=80)
    internal_hash=hash(internal_hash) 
    query="INSERT into images (user_id,folder_id,file_name,internal_hash,description,image_type,content_type,created_at,codebase_version,picture) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title).replace("output/","")+".png"+"',"+str(internal_hash)+",'"+str(description)+"','Kinship','image/png',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=open(Params.output_image+title+'.png','r').read() 
    cur.execute(query,args)
    query="select id from images where internal_hash="+str(internal_hash)+" and image_type like \"%Kinship%\";"
    cur.execute(query) 
    im_id=cur.fetchone()[0]  
    return im_id



def excelOutput(pair_dict,cur,samples,user_id,folder_id,confusionMatrix_degree,header_degree,details_degree,labels_degree,confusionMatrix_rel,details_rel,header_rel,labels_rel,parameter_group,bayes_details,bayes_confusion,training_module,quality,locus_group_name,locus_group_snps): 
    sample_metadata,internal_hash,title,description=get_metadata(samples,cur,training_module,quality,locus_group_name)     
    wb=Workbook() 
    ws=wb.get_active_sheet() 
    ws.title="Kinship Coefficient Details" 
    #add header row 
    headers=['ID1','ID2','marker','IBS0','KinC','mashared','nSharedLoci','x00','x01','x02','x11','x12','x22','relationship','degree','expectedshareddna']
    for i in range(len(headers)): 
        c=ws.cell(row=0,column=i)
        c.value=headers[i]
        print str(c.value) 
    #iterate through all samples and populate the fields 
    index_vals=range(int((len(samples)-1)*len(samples)*.5))
    for j in index_vals: 
        for i in range(len(headers)): 
            c=ws.cell(row=j+1,column=i)
            header=headers[i]
            c.value=pair_dict[header][j] 

    #add a sheet for the degree level confusion matrix 
    #print str(confusionMatrix_degree) 
    #print (confusionMatrix_degree.values)
    ws2=wb.create_sheet() 
    c=ws2.cell(row=0,column=1) 
    c.value="Truth" 
    c=ws2.cell(row=1,column=0) 
    c.value="Predicted"
    labels_degree=list(labels_degree) 
    labels_degree.sort() 

    print str(labels_degree)

    labels_degree_ordered=labels_degree 
    has_unrelated=False 
    if '-1' in labels_degree_ordered: 
        labels_degree_ordered.remove('-1') 
        has_unrelated=True 
    if has_unrelated: 
        labels_degree_ordered.append('-1') 
    print "labels_degree_ordered:"+str(labels_degree_ordered) 
    for i in range(2,len(labels_degree_ordered)+2): 
        c=ws2.cell(row=0,column=i) 
        c.value=labels_degree_ordered[i-2] 
    for j in range(2,len(labels_degree_ordered)+2): 
        c=ws2.cell(row=j,column=0) 
        c.value=labels_degree_ordered[j-2] 
    for i in range(2,len(labels_degree_ordered)+2): 
        for j in range(2,len(labels_degree_ordered)+2): 
            #print "(i,j):("+str(i)+","+str(j)+")"
            c=ws2.cell(row=j,column=i) 
            c.value=confusionMatrix_degree[labels_degree_ordered[j-2]][labels_degree_ordered[i-2]]
    ws2.title="Conf. Mat. (Degree)" 

    #add a sheet for the details of the degree level confusion matrix 
    ws3=wb.create_sheet() 
    ws3.title="Conf. Mat. (Degree) Det."
    for i in range(len(header_degree)): 
        c=ws3.cell(row=0,column=i) 
        c.value=header_degree[i] 
    for j in range(len(details_degree)): 
        for i in range(len(header_degree)): 
            c=ws3.cell(row=j+1,column=i) 
            c.value=details_degree[j][i]  
    
    #add a sheet for the relationship level confusion matrix 
    ws4=wb.create_sheet() 
    c=ws4.cell(row=0,column=1) 
    c.value="Truth" 
    c=ws4.cell(row=1,column=0) 
    c.value="Predicted" 
    labels_rel_ordered=[] 
    labels_rel_candidates=['parent','sibling','grandparent','aunt/uncle','half sibling','great grandparent','great aunt/great uncle','cousin','first cousin','great great grandparent','great great aunt/great great uncle','first cousin once removed','great great great grandparent','great great great aunt/great great great uncle','first cousin twice removed','second cousin','first cousin thrice removed','second cousin once removed','second cousin twice removed','second cousin thrice removed','third cousin','third cousin once removed','third cousin twice removed','third cousin thrice removed','unrelated']
    for candidate in labels_rel_candidates: 
        if candidate in labels_rel: 
            labels_rel_ordered.append(candidate) 
    for i in range(2,len(labels_rel_ordered)+2): 
        c=ws4.cell(row=0,column=i) 
        c.value=labels_rel_ordered[i-2] 
    for j in range(2,len(labels_rel_ordered)+2): 
        c=ws4.cell(row=j,column=0) 
        c.value=labels_rel_ordered[j-2] 
    for i in range(2,len(labels_rel_ordered)+2): 
        for j in range(2,len(labels_rel_ordered)+2): 
            c=ws4.cell(row=j,column=i) 
            c.value=confusionMatrix_rel[labels_rel_ordered[j-2]][labels_rel_ordered[i-2]]
    ws4.title="Conf. Mat. (Relationship)" 

    #add a sheet for the details of the relationship level confusion matrix 
    ws5=wb.create_sheet() 
    ws5.title="Conf. Mat. (Relationship) Det."
    for i in range(len(header_rel)): 
        c=ws5.cell(row=0,column=i) 
        c.value=header_rel[i] 
    for j in range(len(details_rel)): 
        for i in range(len(header_rel)): 
            c=ws5.cell(row=j+1,column=i) 
            c.value=details_rel[j][i]  
    
    #add a sheet for Parameters
    ws6=wb.create_sheet() 
    query="select name from users where id="+str(user_id)+";" 
    cur.execute(query) 
    username=cur.fetchone()
    if username!=None: 
        username=username[0] 
    ws6.title="Parameters" 
    c=ws6.cell(row=0,column=0) 
    c.value=str(parameter_group)+","+str(username) 
    c=ws6.cell(row=1,column=0) 
    c.value="Minimum Reads Per Locus" 
    c=ws6.cell(row=1,column=1) 
    c.value=Params.min_calls 
    
    c=ws6.cell(row=2,column=0) 
    c.value="Reference - MAF for 2 Minor Alleles (lower bound)"
    c=ws6.cell(row=2,column=1) 
    c.value=Params.min_2_ma

    c=ws6.cell(row=3,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (lower bound)"
    c=ws6.cell(row=3,column=1) 
    c.value=Params.min_1_ma 


    c=ws6.cell(row=4,column=0) 
    c.value="Reference - MAF for 1 Minor Allele (upper bound)"
    c=ws6.cell(row=4,column=1) 
    c.value=Params.max_1_ma 

    c=ws6.cell(row=5,column=0) 
    c.value="Reference - MAF for 0 Minor Alleles (upper bound)"
    c=ws6.cell(row=5,column=1) 
    c.value=Params.max_0_ma 

    c=ws6.cell(row=6,column=0) 
    c.value="Strand Bias Cumulative Count (lower bound)"
    c=ws6.cell(row=6,column=1) 
    c.value=Params.strand_bias_thresh

    c=ws6.cell(row=7,column=0) 
    c.value="Ambiguous Bias Cumulative Count (lower bound)"
    c=ws6.cell(row=7,column=1) 
    c.value=Params.ambiguous_thresh 

    c=ws6.cell(row=8,column=0) 
    c.value="Maximum Impossible for Kinship" 
    c=ws6.cell(row=8,column=1) 
    c.value=Params.max_impossible
   

    c=ws6.cell(row=9,column=0) 
    c.value="Quality Threshold" 
    c=ws6.cell(row=9,column=1) 
    c.value=quality 

    c=ws6.cell(row=10,column=0) 
    c.value="Training Module" 
    c=ws6.cell(row=10,column=1) 
    c.value=training_module 

    c=ws6.cell(row=11,column=0) 
    c.value="Locus Group" 
    c=ws6.cell(row=11,column=1) 
    c.value=locus_group_name 
                                                                                                                                                                                                                                         
                                                                                                                        
    #Add a sheet listing all contributors 
    ws7=wb.create_sheet() 
    ws7.title="Contributors" 
    c=ws7.cell(row=0,column=0) 
    c.value="Experiment" 
    c=ws7.cell(row=0,column=1) 
    c.value="Experiment Hash - Quality Cutoff" 
    c=ws7.cell(row=0,column=2) 
    c.value="Barcode" 
    c=ws7.cell(row=0,column=3) 
    c.value="Subject" 
    contributors=description.split('\n') 
    for i in range(len(contributors)): 
        tokens=contributors[i].split(' ') 
        for j in range(len(tokens)): 
            c=ws7.cell(row=i+1,column=j)
            c.value=tokens[j]
    '''
    #Add a worksheet for the Bayesian predictions details 
    ws8=wb.create_sheet() 
    ws8.title="Bayesian Predictions" 
    header=["Reference","Subject","Truth Relationship","Ref. mA/mA", "Ref. mA/MA", "Ref. MA/MA", "Pop. mA/mA", "Pop. mA/MA", "Pop MA/MA"]
    for gen in range(1,Params.smax_degree_bayes+1): 
        header.append("Gen"+str(gen)+" mA/mA") 
        header.append("Gen"+str(gen)+" mA/MA") 
        header.append("Gen"+str(gen)+" MA/MA") 
    header.append("Subj. mA/mA") 
    header.append("Subj. mA/MA") 
    header.append("Subj. MA/MA") 
    header.append("Gen 1 imposs.") 
    for gen in range(1,Params.smax_degree_bayes+1): 
        header.append("Log P(Gen"+str(gen)+")")
    header.append("Log P(Unrel)") 
    header.append("Predicted Relationship") 
    header.append("Correct?") 
    for h in range(len(header)): 
        c=ws8.cell(row=0,column=h) 
        c.value=header[h] 
    for i in range(len(bayes_details)): 
        bdline=bayes_details[i] 
        for j in range(len(bdline)): 
            bdval=bdline[j] 
            c=ws8.cell(row=i+1,column=j) 
            c.value=bdval 

    #Add a worksheet for the Bayesian predictions confusion matrix 
    ws9=wb.create_sheet() 
    ws9.title="Bayesian confusion matrix" 
    for i in range(len(bayes_confusion)): 
        bcline=bayes_confusion[i] 
        for j in range(len(bcline)): 
            bcval=bcline[j] 
            c=ws9.cell(row=i,column=j) 
            c.value=bcval 
    if locus_group_snps!="All": 
        ws10=wb.create_sheet() 
        ws10.title="Locus Group SNPs" 
        c=ws10.cell(row=0,column=0)
        c.value=locus_group_name 
        loc_index=1
        for snp in locus_group_snps: 
            c=ws10.cell(row=1,column=0)
            query="select name from loci where id="+str(snp)+";" 
            cur.execute(query) 
            c.value=cur.fetchone()[0] 
            loc_index+=1 
    '''
    title=title+".xlsx"
    wb.save(Params.output_kinexcel+title) 
    internal_hash=hash(internal_hash) 
    query="INSERT into attachments (user_id,folder_id,file_name,internal_hash,description,file_type,content_type,updated_at,codebase_version,contents) VALUES("+str(user_id)+","+str(folder_id)+",'"+str(title)+"',"+str(internal_hash)+",'"+str(description)+"','Kinship','application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',UTC_TIMESTAMP(),"+str(Params.version)+",%s);"
    #print str(query) 
    args=(open(Params.output_kinexcel+title,'r').read(),) 
    cur.execute(query,args)
    query="select id from attachments where internal_hash='"+str(internal_hash)+"';"
    cur.execute(query) 
    attachment_id=cur.fetchone()[0] 
    return attachment_id 
